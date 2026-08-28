#!/usr/bin/env python3
"""Build the Tables S1–S17 workbook from the supplied CSV files."""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
from pathlib import Path
import re
import tempfile
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class TableSpec:
    sheet: str
    title: str
    filename: str
    purpose: str
    note: str


TABLES = (
    TableSpec("S1 Sources", "Sources reviewed and their use in the study", "Table_S1_Sources_Reviewed.csv", "Literature review", "Reviewed sources and their role in the study."),
    TableSpec("S2 Searches", "Literature searches and citation tracing", "Table_S2_Searches_and_Citation_Tracing.csv", "Literature review", "Searches and citation relationships used to locate publications."),
    TableSpec("S3 Citations", "Citation relationships reviewed", "Table_S3_Citation_Relationships.csv", "Literature review", "Links followed from citing publications to reports of disease measurements."),
    TableSpec("S4 Source files", "Source files and checksums", "Table_S4_Source_Files_and_Checksums.csv", "Source documentation", "Public locations, access dates, and SHA-256 values for acquired files."),
    TableSpec("S5 Reused measurements", "Reports based on the same phenotype measurements", "Table_S5_Reports_Using_the_Same_Measurements.csv", "Source documentation", "The table identifies publications that report the same underlying measurements."),
    TableSpec("S6 Studies", "Experiments and assays included in quantitative summaries", "Table_S6_Experiments_and_Assays.csv", "Study description", "Distinct experiments and assays contributing quantitative summaries."),
    TableSpec("S7 Endpoints", "Observation counts for each disease endpoint", "Table_S7_Disease_Endpoint_Counts.csv", "Study description", "Counts by study, disease, endpoint, environment, scale, and germplasm role."),
    TableSpec("S8 Pareto fronts", "Accessions or lines on Pareto fronts within individual studies", "Table_S8_Pareto_Front_Summary.csv", "Secondary analysis", "An accession or line is on the Pareto front when no other accession or line performs at least as well on every endpoint and better on at least one."),
    TableSpec("S9 Candidate criteria", "Criteria used to identify candidates", "Table_S9_Candidate_Criteria.csv", "Candidate definitions", "Documented criteria applied to each experiment or assay."),
    TableSpec("S10 Repeated findings", "Frequency of candidate identification across independent screens", "Table_S10_Repeated_Identification_Results.csv", "Primary analysis", "Observed and permutation results for repeated candidate identification."),
    TableSpec("S11 Study questions", "Numerical results for the two study questions", "Table_S11_Study_Question_Results.csv", "Results summary", "Study-specific disease overlap and repeated candidate identification."),
    TableSpec("S12 Thresholds", "Results obtained with alternative disease-score thresholds", "Table_S12_Threshold_Sensitivity.csv", "Sensitivity analysis", "Puerto Rico disease means classified at four prespecified cutoffs."),
    TableSpec("S12a Disease overlap", "Disease overlap expected from the candidate counts in each study", "Table_S12a_Disease_Overlap_Distribution.csv", "Primary analysis", "Exact results preserving the candidate count for each disease within each study."),
    TableSpec("S13 Reported results", "Sources and calculations for numerical values reported in the manuscript", "Table_S13_Reported_Results_and_Sources.csv", "Reproducibility", "The table gives the source and calculation for each numerical result."),
    TableSpec("S14 Publications", "Publications reviewed from the Prom bibliography", "Table_S14_Publications_Reviewed.csv", "Literature review", "Publication citations and extent of review within the defined citation network."),
    TableSpec("S15 Source review", "Sources found through one round of citation tracing", "Table_S15_Citation_Tracing_Sources.csv", "Literature review", "Availability, relevance, source relationships, and rights information."),
    TableSpec("S16 Identifier searches", "Identifier searches for references without recoverable titles", "Table_S16_Unresolved_References.csv", "Literature review", "Citation strings and identifier searches recorded on 14 July 2026."),
    TableSpec("S17 Repeated candidates", "Accessions or lines identified in at least two independent screens", "Table_S17_Repeated_Candidates.csv", "Summary of repeatedly identified candidates", "The 171 candidates, diseases, qualifying screens, and source citations."),
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.reader(stream)
        try:
            header = next(reader)
        except StopIteration as exc:
            raise ValueError(f"Empty CSV: {path}") from exc
        rows = [row for row in reader]
    if len(set(header)) != len(header):
        raise ValueError(f"Duplicate headers in {path.name}")
    if any(len(row) != len(header) for row in rows):
        raise ValueError(f"Rows with inconsistent column counts in {path.name}")
    return header, rows


def scalar(value: str, header: str) -> object:
    stripped = value.strip()
    if stripped == "":
        return None
    numeric_header = (
        header.endswith("(n)")
        or header
        in {
            "Review version",
            "Reference depth",
            "Results returned",
            "Results screened",
            "Download byte size",
            "Archive member byte size",
            "Confidence",
            "Disease count",
            "Candidates among those assessed",
            "Random seed",
            "Puerto Rico disease-score cutoff",
            "Numeric result",
            "Reported text",
            "Year",
            "Source identifiers in Table S1",
            "Max reported reference count",
            "Reference count captured",
            "Citation-tracing depth",
            "Number of qualifying independent screens",
        }
        or any(
            phrase in header.lower()
            for phrase in (
                "proportion",
                "probability",
                "exact expected mean",
                "expected median",
                "lower limit",
                "upper limit",
                "permutations",
                "random seed",
            )
        )
    )
    if not numeric_header:
        return value
    try:
        if stripped.lstrip("+-").isdigit():
            return int(stripped)
        return float(stripped)
    except ValueError:
        return value


def style_sheet(worksheet, row_count: int, column_count: int) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}{row_count + 1}"
    worksheet.print_title_rows = "1:1"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_LETTER
    worksheet.page_setup.fitToWidth = 2 if column_count > 18 else 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_margins.left = 0.35
    worksheet.page_margins.right = 0.35
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5
    worksheet.row_dimensions[1].height = 36

    header_fill = PatternFill("solid", fgColor="285943")
    band_fill = PatternFill("solid", fgColor="EAF5F8")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    header_border = Border(bottom=Side(style="thin", color="1D4434"))
    body_border = Border(bottom=Side(style="hair", color="D9E2E6"))

    for cell in worksheet[1]:
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = header_border

    for row_index in range(2, row_count + 2):
        fill = band_fill if row_index % 2 == 0 else white_fill
        for cell in worksheet[row_index]:
            cell.font = Font(name="Arial", size=9, color="1F2933")
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = body_border

    sample_limit = min(row_count + 1, 202)
    for column_index in range(1, column_count + 1):
        values = [
            str(worksheet.cell(row, column_index).value or "")
            for row in range(1, sample_limit + 1)
        ]
        width = min(max(max((len(value) for value in values), default=8) + 2, 12), 48)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def normalize_xlsx(source: Path, output: Path) -> None:
    timestamp = (2026, 8, 25, 0, 0, 0)
    with ZipFile(source, "r") as input_zip, ZipFile(
        output, "w", compression=ZIP_DEFLATED, compresslevel=9
    ) as output_zip:
        for name in sorted(input_zip.namelist()):
            info = ZipInfo(name, timestamp)
            info.compress_type = ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            payload = input_zip.read(name)
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2026-08-25T00:00:00Z\g<2>",
                    payload,
                )
            output_zip.writestr(info, payload)


def build(tables_dir: Path, output: Path) -> dict[str, object]:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "Sorghum disease-screen supplementary tables"
    workbook.properties.creator = "James Liu; Louis K. Prom; Ezekiel Ahn"
    fixed_time = datetime(2026, 8, 25, tzinfo=timezone.utc).replace(tzinfo=None)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time

    readme = workbook.create_sheet("README")
    readme.append(["Table", "Title", "Source CSV", "Rows", "SHA-256", "Purpose", "Notes"])

    reports: list[dict[str, object]] = []
    for spec in TABLES:
        source = tables_dir / spec.filename
        if not source.is_file():
            raise FileNotFoundError(source)
        header, text_rows = read_csv(source)
        worksheet = workbook.create_sheet(spec.sheet)
        worksheet.append(header)
        for row in text_rows:
            worksheet.append(
                [scalar(value, header[index]) for index, value in enumerate(row)]
            )
        style_sheet(worksheet, len(text_rows), len(header))
        file_hash = sha256(source)
        readme.append(
            [spec.sheet, spec.title, f"supplementary_tables/{spec.filename}", len(text_rows), file_hash, spec.purpose, spec.note]
        )
        reports.append({"sheet": spec.sheet, "rows": len(text_rows), "sha256": file_hash})

    style_sheet(readme, len(TABLES), 7)
    for column, width in {"A": 25, "B": 42, "C": 58, "D": 14, "E": 68, "F": 28, "G": 72}.items():
        readme.column_dimensions[column].width = width
    for row_index in range(2, len(TABLES) + 2):
        readme.row_dimensions[row_index].height = 45

    with tempfile.TemporaryDirectory(prefix="supplementary_workbook_") as temporary:
        raw = Path(temporary) / "workbook.xlsx"
        workbook.save(raw)
        output.parent.mkdir(parents=True, exist_ok=True)
        normalize_xlsx(raw, output)

    check = load_workbook(output, read_only=True, data_only=True)
    expected_sheets = ["README", *(spec.sheet for spec in TABLES)]
    if check.sheetnames != expected_sheets:
        raise ValueError("Worksheet order changed after saving")
    if check["S12a Disease overlap"].max_row != 7:
        raise ValueError("Table S12a must contain six results")
    registry = check["S17 Repeated candidates"]
    if registry.max_row != 172:
        raise ValueError("Table S17 must contain 171 candidates")
    header = [cell.value for cell in next(registry.iter_rows(min_row=1, max_row=1))]
    pattern_column = header.index("Repeated identification pattern")
    patterns = [row[pattern_column] for row in registry.iter_rows(min_row=2, values_only=True)]
    if patterns.count("Same disease") != 78 or patterns.count("Different diseases only") != 93:
        raise ValueError(
            "Table S17 must contain 78 same-disease and 93 different-disease candidates"
        )
    check.close()

    return {
        "validation": "PASS",
        "output": str(output),
        "sha256": sha256(output),
        "worksheets": 19,
        "supplementary_tables": 18,
        "repeated_candidates": 171,
        "same_disease": 78,
        "different_diseases_only": 93,
        "sources": reports,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--tables-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    report = build(args.tables_dir.resolve(), args.output.resolve())
    print(
        "PASS: wrote "
        f"{report['supplementary_tables']} supplementary tables to {report['output']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
