#!/usr/bin/env python3
"""Validate the released inputs and reproduced headline results."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd
from PIL import Image


EXPECTED_EXACT = {
    "recurrence_evidence_set_n": 34,
    "recurrence_assessed_identity_n": 1154,
    "recurrence_candidate_n": 612,
    "recurrence_identified_once_n": 441,
    "recurrence_identified_repeatedly_n": 171,
    "fixed_margin_block_n": 5,
    "fixed_margin_complete_case_n": 556,
    "fixed_margin_observed_multidisease_n": 115,
    "fixed_margin_null_median": 115,
    "fixed_margin_null_lower_95": 105,
    "fixed_margin_null_upper_95": 125,
}


def close(actual: float, expected: float, tolerance: float = 1e-12) -> bool:
    return math.isclose(float(actual), float(expected), rel_tol=tolerance, abs_tol=tolerance)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def validate(
    recurrence_input: Path,
    fixed_margin_input: Path,
    output_dir: Path,
    registry_path: Path | None = None,
    registry_memberships_path: Path | None = None,
    package_root: Path | None = None,
) -> dict[str, object]:
    recurrence = pd.read_csv(output_dir / "Repeated_Identification_Summary.csv").iloc[0]
    fixed = pd.read_csv(output_dir / "Disease_Overlap_Summary.csv")
    fixed_combined = fixed.loc[
        fixed["Comparison identifier"].eq(
            "declared_multidisease_comparisons_fixed_margin"
        )
    ].iloc[0]
    recurrence_matrix = pd.read_csv(recurrence_input)
    fixed_input = pd.read_csv(fixed_margin_input)

    observed = {
        "recurrence_evidence_set_n": int(recurrence["Independent screens (n)"]),
        "recurrence_assessed_identity_n": int(
            recurrence["Assessed in at least two screens (n)"]
        ),
        "recurrence_candidate_n": int(
            recurrence["Candidates assessed in at least two screens (n)"]
        ),
        "recurrence_identified_once_n": int(
            recurrence["Candidates identified once (n)"]
        ),
        "recurrence_identified_repeatedly_n": int(
            recurrence["Candidates identified repeatedly (n)"]
        ),
        "fixed_margin_block_n": len(fixed_input),
        "fixed_margin_complete_case_n": int(fixed_combined["Accessions or lines assessed (n)"]),
        "fixed_margin_observed_multidisease_n": int(
            fixed_combined["Observed multidisease classifications (n)"]
        ),
        "fixed_margin_null_median": int(fixed_combined["Expected median"]),
        "fixed_margin_null_lower_95": int(
            fixed_combined["Central 95% interval, lower limit"]
        ),
        "fixed_margin_null_upper_95": int(
            fixed_combined["Central 95% interval, upper limit"]
        ),
    }
    failures = [
        f"{name}: expected {expected}, found {observed[name]}"
        for name, expected in EXPECTED_EXACT.items()
        if observed[name] != expected
    ]

    approximate_checks = {
        "recurrence_observed_fraction": (
            float(recurrence["Observed proportion identified repeatedly"]),
            171 / 612,
            1e-12,
        ),
        "recurrence_permutation_median": (
            float(recurrence["Permutation median"]),
            0.2147239263803681,
            1e-12,
        ),
        "recurrence_permutation_lower_95": (
            float(recurrence["Central 95% permutation interval, lower limit"]),
            0.1914893617021276,
            1e-12,
        ),
        "recurrence_permutation_upper_95": (
            float(recurrence["Central 95% permutation interval, upper limit"]),
            0.2380952380952380,
            1e-12,
        ),
        "recurrence_two_sided_empirical_p": (
            float(recurrence["Two-sided empirical P"]),
            0.0001999800019998,
            1e-12,
        ),
        "fixed_margin_exact_mean": (
            float(fixed_combined["Exact expected mean"]),
            114.9840799672634,
            1e-12,
        ),
        "fixed_margin_lower_tail": (
            float(fixed_combined["Lower-tail probability"]),
            0.5401614321764678,
            1e-12,
        ),
        "fixed_margin_upper_tail": (
            float(fixed_combined["Upper-tail probability"]),
            0.5372078142877608,
            1e-12,
        ),
    }
    for name, (actual, expected, tolerance) in approximate_checks.items():
        if not close(actual, expected, tolerance):
            failures.append(
                f"{name}: expected {expected} within {tolerance}, found {actual}"
            )

    if recurrence_matrix["analysis_identity_id"].astype(str).str.match(
        r"^(PI\s*\d+|pi\d+|line:)", case=False
    ).any():
        failures.append(
            "The anonymized table used to analyze repeated identification contains a recognizable accession or line identifier."
        )

    registry_checks: dict[str, object] = {"registry_checked": False}
    if registry_path is not None and registry_path.exists():
        registry = pd.read_csv(registry_path, dtype=str, keep_default_na=False)
        required = {
            "normalized_accession_or_line_id",
            "recurrence_pattern",
            "qualifying_evidence_set_n",
            "disease_categories",
            "source_citations",
        }
        missing = sorted(required - set(registry.columns))
        patterns = registry.get("recurrence_pattern", pd.Series(dtype=str)).value_counts()
        registry_checks = {
            "registry_checked": True,
            "registry_rows": len(registry),
            "registry_same_disease_recurrence_n": int(
                patterns.get("same_disease_recurrence", 0)
            ),
            "registry_cross_disease_only_n": int(patterns.get("cross_disease_only", 0)),
            "registry_missing_required_columns": missing,
            "registry_duplicate_normalized_ids": int(
                registry.get("normalized_accession_or_line_id", pd.Series(dtype=str))
                .duplicated()
                .sum()
            ),
        }
        if len(registry) != 171:
            failures.append(f"Registry expected 171 candidates, found {len(registry)}")
        if int(patterns.get("same_disease_recurrence", 0)) != 78:
            failures.append(
                "Expected 78 candidates identified repeatedly for the same disease"
            )
        if int(patterns.get("cross_disease_only", 0)) != 93:
            failures.append(
                "Expected 93 candidates identified only across different diseases"
            )
        if missing:
            failures.append(f"Registry missing required columns: {missing}")
        if registry_checks["registry_duplicate_normalized_ids"]:
            failures.append("Registry contains duplicate normalized identifiers")
        if not missing and registry[list(required)].apply(
            lambda column: column.astype(str).str.strip().eq("")
        ).any().any():
            failures.append("Registry contains blank required values")
        evidence_counts = pd.to_numeric(
            registry.get("qualifying_evidence_set_n", pd.Series(dtype=str)),
            errors="coerce",
        ).value_counts()
        registry_checks["registry_evidence_set_count_distribution"] = {
            str(int(key)): int(value) for key, value in evidence_counts.items()
        }
        if registry_checks["registry_evidence_set_count_distribution"] != {
            "2": 164,
            "3": 7,
        }:
            failures.append(
                "Expected 164 candidates in two independent screens and 7 in three independent screens"
            )
        forbidden_name_parts = (
            "phenotype",
            "score",
            "severity",
            "mean",
            "ranking",
            "threshold",
        )
        forbidden_columns = [
            column
            for column in registry.columns
            if any(part in column.lower() for part in forbidden_name_parts)
        ]
        registry_checks["registry_forbidden_measurement_columns"] = forbidden_columns
        if forbidden_columns:
            failures.append(
                f"Registry contains prohibited measurement columns: {forbidden_columns}"
            )

    if registry_memberships_path is not None and registry_memberships_path.exists():
        memberships = pd.read_csv(
            registry_memberships_path, dtype=str, keep_default_na=False
        )
        registry_checks["registry_membership_rows"] = len(memberships)
        if len(memberships) != 401:
            failures.append(
                f"Registry expected 401 qualifying observations, found {len(memberships)}"
            )
        unique_candidate_screen_links = memberships[
            ["normalized_accession_or_line_id", "evidence_set_id"]
        ].drop_duplicates()
        registry_checks["registry_unique_candidate_screen_links"] = len(
            unique_candidate_screen_links
        )
        if len(unique_candidate_screen_links) != 349:
            failures.append(
                "Expected 349 distinct combinations of candidates and independent screens, "
                f"found {len(unique_candidate_screen_links)}"
            )
        if registry_path is None or not registry_path.exists():
            failures.append(
                "Observations of qualifying candidates require the detailed registry"
            )
        else:
            membership_ids = set(memberships["normalized_accession_or_line_id"])
            registry_ids = set(registry["normalized_accession_or_line_id"])
            if membership_ids != registry_ids:
                failures.append(
                    "Candidate IDs differ between the detailed registry and observations of qualifying candidates"
                )
            counts = memberships.groupby("normalized_accession_or_line_id")[
                "evidence_set_id"
            ].nunique()
            declared = registry.set_index("normalized_accession_or_line_id")[
                "qualifying_evidence_set_n"
            ].astype(int)
            if not counts.sort_index().equals(declared.sort_index()):
                failures.append(
                    "Counts of independent screens differ between the registry and observations of qualifying candidates"
                )

    package_checks: dict[str, object] = {"package_checked": False}
    if package_root is not None:
        package_root = package_root.resolve()
        required_files = [
            "figures/Figure_1.tif",
            "figures/Figure_2.tif",
            "figures/Figure_3.tif",
            "supplement/Figure_S1.pdf",
            "supplement/Supplementary_Tables_S1-S17.xlsx",
            "supplement/Recurrent_Candidate_Registry.csv",
            "reproducibility/code/01_data_overview.ipynb",
            "reproducibility/code/02_manuscript_analyses.ipynb",
        ]
        absent = [name for name in required_files if not (package_root / name).is_file()]
        if absent:
            failures.append(f"Required package files are absent: {absent}")

        unwanted = []
        for path in package_root.rglob("*"):
            if not path.is_file():
                continue
            relative_path = path.relative_to(package_root).as_posix()
            is_unwanted_markdown = (
                path.suffix.lower() in {".md", ".markdown"}
                and relative_path != "README.md"
            )
            if (
                is_unwanted_markdown
                or path.suffix.lower() == ".pyc"
                or path.name == ".DS_Store"
            ):
                unwanted.append(relative_path)
        if unwanted:
            failures.append(f"Unwanted package files found: {unwanted}")

        local_path_hits: list[str] = []
        for path in package_root.rglob("*"):
            if not path.is_file() or path.suffix.lower() not in {
                ".txt", ".md", ".csv", ".json", ".py", ".ipynb"
            }:
                continue
            text = path.read_text(encoding="utf-8", errors="ignore")
            local_tokens = (
                "/" + "Users" + "/",
                "file" + "://",
                ".local" + "-only",
            )
            if any(token in text for token in local_tokens):
                local_path_hits.append(str(path.relative_to(package_root)))
        if local_path_hits:
            failures.append(f"Local-host paths found in: {local_path_hits}")

        data_root = package_root / "reproducibility" / "data"
        dictionary_path = data_root / "DATA_DICTIONARY.csv"
        dictionary_checks: dict[str, object] = {}
        if dictionary_path.is_file():
            dictionary = pd.read_csv(dictionary_path, dtype=str, keep_default_na=False)
            listed_files = set(dictionary["File path"])
            supplied_files = {
                str(path.relative_to(data_root))
                for path in data_root.rglob("*")
                if path.is_file()
            }
            dictionary_checks = {
                "listed_files": len(listed_files),
                "supplied_files": len(supplied_files),
                "missing_entries": sorted(supplied_files - listed_files),
                "absent_files": sorted(listed_files - supplied_files),
            }
            if dictionary_checks["missing_entries"] or dictionary_checks["absent_files"]:
                failures.append(f"Data dictionary and supplied files disagree: {dictionary_checks}")
        else:
            failures.append("The data dictionary is absent")

        source_manifest_path = data_root / "source_metadata" / "source_snapshot_manifest.json"
        source_table_path = data_root / "supplementary_tables" / "Table_S4_Source_Files_and_Checksums.csv"
        source_checks: dict[str, object] = {}
        if source_manifest_path.is_file() and source_table_path.is_file():
            source_manifest = json.loads(source_manifest_path.read_text(encoding="utf-8"))
            source_table = pd.read_csv(source_table_path, dtype=str, keep_default_na=False)
            source_checks = {
                "manifest_records": len(source_manifest["records"]),
                "declared_records": int(source_manifest["record_count"]),
                "source_table_rows": len(source_table),
                "recorded_checksums": int(source_manifest["recorded_checksum_count"]),
                "source_table_sha256_matches": source_manifest["source_file_table_sha256"]
                == sha256(source_table_path),
            }
            if source_checks != {
                "manifest_records": 178,
                "declared_records": 178,
                "source_table_rows": 178,
                "recorded_checksums": 100,
                "source_table_sha256_matches": True,
            }:
                failures.append(f"Source documentation changed: {source_checks}")
        else:
            failures.append("Source documentation is incomplete")

        aggregate_root = data_root / "aggregate_data"
        aggregate_checks: dict[str, object] = {}
        checksum_path = aggregate_root / "AGGREGATE_OUTPUT_CHECKSUMS.txt"
        output_manifest_path = aggregate_root / "CURRENT_OUTPUT_MANIFEST.json"
        if checksum_path.is_file() and output_manifest_path.is_file():
            declared_hashes = {}
            for line in checksum_path.read_text(encoding="utf-8").splitlines():
                digest, filename = line.split(maxsplit=1)
                declared_hashes[filename.strip()] = digest
            output_manifest = json.loads(output_manifest_path.read_text(encoding="utf-8"))
            current_hashes = {
                entry["name"]: sha256(aggregate_root / entry["name"])
                for entry in output_manifest["files"]
            }
            aggregate_checks = {
                "result_files": len(current_hashes),
                "checksum_list_matches": declared_hashes == current_hashes,
                "manifest_hashes_match": all(
                    entry["sha256"] == current_hashes[entry["name"]]
                    for entry in output_manifest["files"]
                ),
                "manifest_sizes_match": all(
                    int(entry["bytes"]) == (aggregate_root / entry["name"]).stat().st_size
                    for entry in output_manifest["files"]
                ),
            }
            if aggregate_checks != {
                "result_files": 5,
                "checksum_list_matches": True,
                "manifest_hashes_match": True,
                "manifest_sizes_match": True,
            }:
                failures.append(f"Supplied result-file documentation changed: {aggregate_checks}")
        else:
            failures.append("Result-file checksums or manifest are absent")

        figure_expectations = {
            "Figure_1.tif": (4200, 4620),
            "Figure_2.tif": (4200, 2550),
            "Figure_3.tif": (4200, 2550),
        }
        figure_checks: dict[str, object] = {}
        for filename, expected_pixels in figure_expectations.items():
            path = package_root / "figures" / filename
            if not path.is_file():
                continue
            with Image.open(path) as image:
                details = {
                    "pixels": list(image.size),
                    "mode": image.mode,
                    "dpi": [float(value) for value in image.info.get("dpi", ())],
                    "compression_tag": int(image.tag_v2.get(259, -1)),
                }
            figure_checks[filename] = details
            if tuple(details["pixels"]) != expected_pixels:
                failures.append(f"{filename} pixel dimensions changed")
            if details["mode"] != "RGB":
                failures.append(f"{filename} must be RGB")
            if details["compression_tag"] != 5:
                failures.append(f"{filename} must use LZW compression")
            if not details["dpi"] or any(abs(float(value) - 600.0) > 0.01 for value in details["dpi"]):
                failures.append(f"{filename} must record 600-dpi resolution")

        supplementary_pdf = package_root / "supplement" / "Figure_S1.pdf"
        if supplementary_pdf.is_file():
            pdf_bytes = supplementary_pdf.read_bytes()
            if b"Arial" not in pdf_bytes or b"DejaVu" in pdf_bytes:
                failures.append("Figure S1 must embed Arial and exclude DejaVu Sans")

        workbook_path = package_root / "supplement" / "Supplementary_Tables_S1-S17.xlsx"
        workbook_checks: dict[str, object] = {}
        if workbook_path.is_file():
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            workbook_checks = {
                "worksheets": len(workbook.sheetnames),
                "table_s12a_rows": workbook["S12a Disease overlap"].max_row - 1,
                "table_s17_rows": workbook["S17 Repeated candidates"].max_row - 1,
            }
            if workbook_checks != {
                "worksheets": 19,
                "table_s12a_rows": 6,
                "table_s17_rows": 171,
            }:
                failures.append(f"Supplementary workbook counts changed: {workbook_checks}")
            workbook.close()

        package_checks = {
            "package_checked": True,
            "required_files_absent": absent,
            "unwanted_files": unwanted,
            "local_path_hits": local_path_hits,
            "data_dictionary": dictionary_checks,
            "source_documentation": source_checks,
            "aggregate_results": aggregate_checks,
            "figures": figure_checks,
            "supplementary_workbook": workbook_checks,
        }

    result: dict[str, object] = {
        "validation_status": "PASS" if not failures else "FAIL",
        "failures": failures,
        "headline_results": observed,
        "anonymous_recurrence_input_rows": len(recurrence_matrix),
        **registry_checks,
        **package_checks,
    }
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--recurrence-input", type=Path, required=True)
    parser.add_argument("--fixed-margin-input", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--registry", type=Path)
    parser.add_argument("--registry-memberships", type=Path)
    parser.add_argument("--package-root", type=Path)
    args = parser.parse_args()
    result = validate(
        args.recurrence_input,
        args.fixed_margin_input,
        args.output_dir,
        args.registry,
        args.registry_memberships,
        args.package_root,
    )
    (args.output_dir / "analysis_validation.json").write_text(
        json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    print(json.dumps(result, sort_keys=True))
    return 0 if result["validation_status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
